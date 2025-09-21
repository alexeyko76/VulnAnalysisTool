#!/usr/bin/env python3
"""
Script to specifically check Weblogic CVEs in the CVE sheet
"""

import openpyxl
from openpyxl import load_workbook

def check_weblogic_cves():
    excel_file = "./sample-data/sample.xlsx"

    try:
        wb = load_workbook(excel_file)
        cve_sheet = wb["CVEs"]

        print("=== WEBLOGIC CVE DETECTION TEST ===\n")

        # Check all rows for Weblogic-related CVEs
        weblogic_cves = []
        oracle_advisories_found = []

        for row_num in range(2, cve_sheet.max_row + 1):
            row = cve_sheet[row_num]
            cve_id = row[0].value if row[0].value else ""
            description = row[1].value if row[1].value else ""
            references = row[2].value if row[2].value else ""
            affected_software = row[3].value if row[3].value else ""
            weblogic_flag = row[4].value if row[4].value else ""
            oracle_advisories = row[5].value if row[5].value else ""

            # Check for known Weblogic CVEs
            if cve_id in ["CVE-2020-14882", "CVE-2017-10271"]:
                weblogic_cves.append({
                    'cve_id': cve_id,
                    'description': description[:100] + "..." if len(str(description)) > 100 else description,
                    'weblogic_flag': weblogic_flag,
                    'affected_software': affected_software,
                    'oracle_advisories': oracle_advisories
                })

            # Check for any Oracle advisories
            if oracle_advisories and "oracle.com" in str(oracle_advisories):
                oracle_advisories_found.append(cve_id)

        print("WEBLOGIC-SPECIFIC CVEs:")
        if weblogic_cves:
            for cve in weblogic_cves:
                print(f"  {cve['cve_id']}:")
                print(f"    Description: {cve['description']}")
                print(f"    Weblogic Flag: {cve['weblogic_flag']}")
                print(f"    Affected Software: {cve['affected_software'][:200]}...")
                print(f"    Oracle Advisories: {cve['oracle_advisories']}")
                print()
        else:
            print("  No Weblogic CVEs found in the sheet")

        print(f"ORACLE ADVISORIES FOUND: {len(oracle_advisories_found)}")
        for cve_id in oracle_advisories_found:
            print(f"  {cve_id}")

        # Show total statistics
        total_rows = cve_sheet.max_row - 1  # Exclude header
        weblogic_y_count = 0
        oracle_advisory_count = 0

        for row_num in range(2, cve_sheet.max_row + 1):
            row = cve_sheet[row_num]
            weblogic_flag = row[4].value if row[4].value else ""
            oracle_advisories = row[5].value if row[5].value else ""

            if weblogic_flag == "Y":
                weblogic_y_count += 1
            if oracle_advisories and "oracle.com" in str(oracle_advisories):
                oracle_advisory_count += 1

        print(f"\nSUMMARY:")
        print(f"  Total CVEs processed: {total_rows}")
        print(f"  CVEs flagged as Weblogic: {weblogic_y_count}")
        print(f"  CVEs with Oracle advisories: {oracle_advisory_count}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    check_weblogic_cves()