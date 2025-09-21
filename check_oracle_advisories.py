#!/usr/bin/env python3
"""
Script to check Oracle advisories extraction for specific CVEs
"""

import openpyxl
from openpyxl import load_workbook

def check_oracle_advisories():
    excel_file = "./sample-data/sample.xlsx"

    try:
        wb = load_workbook(excel_file)
        cve_sheet = wb["CVEs"]

        print("=== ORACLE ADVISORIES ANALYSIS ===\n")

        # Check specific CVEs that should have Oracle advisories
        target_cves = ["CVE-2020-14882", "CVE-2017-10271", "CVE-2024-20931", "CVE-2023-21839"]

        for row_num in range(2, cve_sheet.max_row + 1):
            row = cve_sheet[row_num]
            cve_id = row[0].value if row[0].value else ""

            if cve_id in target_cves:
                description = row[1].value if row[1].value else ""
                references = row[2].value if row[2].value else ""
                affected_software = row[3].value if row[3].value else ""
                weblogic_flag = row[4].value if row[4].value else ""
                oracle_advisories = row[5].value if row[5].value else ""

                print(f"{cve_id}:")
                print(f"  Weblogic Flag: {weblogic_flag}")
                print(f"  Oracle Advisories: '{oracle_advisories}'")
                print(f"  References (first 500 chars):")
                if references:
                    refs_str = str(references)
                    print(f"    {refs_str[:500]}...")

                    # Look for Oracle URLs manually
                    oracle_urls = []
                    if "oracle.com" in refs_str.lower():
                        # Split by common delimiters and look for Oracle URLs
                        parts = refs_str.replace(';', '\n').replace(',', '\n').split('\n')
                        for part in parts:
                            part = part.strip()
                            if 'oracle.com' in part.lower():
                                oracle_urls.append(part)

                    print(f"  Found Oracle URLs manually: {len(oracle_urls)}")
                    for url in oracle_urls[:3]:  # Show first 3
                        print(f"    {url}")
                else:
                    print("    No references found")
                print()

        # Also check for any CVE that has Oracle advisories
        print("=== CVEs WITH ORACLE ADVISORIES ===")
        oracle_advisory_count = 0
        for row_num in range(2, cve_sheet.max_row + 1):
            row = cve_sheet[row_num]
            cve_id = row[0].value if row[0].value else ""
            oracle_advisories = row[5].value if row[5].value else ""

            if oracle_advisories and oracle_advisories.strip():
                oracle_advisory_count += 1
                print(f"{cve_id}: {oracle_advisories}")

        print(f"\nTotal CVEs with Oracle advisories: {oracle_advisory_count}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    check_oracle_advisories()