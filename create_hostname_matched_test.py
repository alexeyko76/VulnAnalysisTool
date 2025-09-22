#!/usr/bin/env python3
"""
Create test data that matches current hostname for actual testing
"""

import openpyxl
from openpyxl import Workbook
import os
import socket

def create_hostname_matched_test():
    # Get current hostname (same as Java tool will get)
    try:
        current_hostname = socket.gethostname()
        print(f"Current hostname: {current_hostname}")
    except:
        current_hostname = "LPRIME"  # Fallback

    # Create new test data
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Export"

    # Headers
    headers = ["AFFECTED_PLATFORMS", "XTRACT_PATH", "HOSTNAME", "CVE"]
    for i, header in enumerate(headers, 1):
        sheet.cell(row=1, column=i, value=header)

    # Get absolute paths to our test files
    base_path = os.path.abspath("test-linux-files")

    # Test data using current hostname but Linux platform
    # This simulates running the tool on a Linux machine with the current hostname
    test_data = [
        # JAR files - should be processed for version extraction
        ["Linux", f"{base_path}/usr/lib/jvm/java-8-openjdk/jre/lib/rt.jar", current_hostname, "CVE-2022-1234"],
        ["Linux", f"{base_path}/opt/tomcat/lib/catalina.jar", current_hostname, "CVE-2022-5678"],
        ["Linux", f"{base_path}/usr/share/java/log4j-core-2.14.1.jar", current_hostname, "CVE-2021-44228"],

        # EXE files - should be skipped for version processing because platform is Linux
        ["Linux", f"{base_path}/home/user/.wine/drive_c/Program Files/WinApp/app.exe", current_hostname, "CVE-2022-9999"],

        # Other file types
        ["Linux", f"{base_path}/usr/bin/curl", current_hostname, "CVE-2022-2222"],

        # Non-existent file
        ["Linux", f"{base_path}/invalid/path/does/not/exist.jar", current_hostname, "CVE-2022-0000"],

        # Invalid paths
        ["Linux", "", current_hostname, "CVE-2022-3333"],
        ["Linux", "N/A", current_hostname, "CVE-2022-4444"],

        # Corrupted path (should trigger path fixing)
        ["Linux", f"{base_path}/usr/share/java/log4j-core-2.14.1.jar extra_garbage_data", current_hostname, "CVE-2022-5555"],
    ]

    for i, row_data in enumerate(test_data, 2):
        for j, value in enumerate(row_data, 1):
            sheet.cell(row=i, column=j, value=value)

    # Save test file
    test_file = "./sample-data/linux-hostname-test.xlsx"
    wb.save(test_file)
    print(f"Created hostname-matched Linux test file: {test_file}")

    # Create corresponding config
    config_content = f"""# Linux test with current hostname
excel.path=./sample-data/linux-hostname-test.xlsx
sheet.name=Export
column.PlatformName=AFFECTED_PLATFORMS
column.FilePath=XTRACT_PATH
column.HostName=HOSTNAME
column.CVE=CVE
platform.windows=Windows Server 2016, Windows Server 2019, Windows Server 2022
remote.unc.enabled=false
remote.unc.timeout=6
log.filename=linux-hostname-test.log
invalid.path.detection=true
duplicate.search.enabled=false
cve.sheet.creation.enabled=false
"""

    with open("linux-hostname-test-config.properties", "w") as f:
        f.write(config_content)

    print("Created config: linux-hostname-test-config.properties")

    # Show test summary
    print(f"\nTest summary for hostname '{current_hostname}' with Linux platform:")
    print("- 3 JAR files (should extract versions)")
    print("- 1 EXE file (should skip version extraction - Linux platform)")
    print("- 1 other file (should check existence only)")
    print("- 1 non-existent file (should mark as N)")
    print("- 2 invalid paths (should mark as X)")
    print("- 1 corrupted path (should attempt fixing)")
    print()
    print("This will test Linux logic while running on current system!")

if __name__ == "__main__":
    create_hostname_matched_test()