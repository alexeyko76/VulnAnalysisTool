#!/usr/bin/env python3
"""
Script to add real Weblogic CVEs from 2020-2024 for testing Weblogic detection
"""

import openpyxl
from openpyxl import load_workbook
import os

def add_real_weblogic_cves():
    excel_file = "./sample-data/sample.xlsx"

    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found")
        return

    try:
        wb = load_workbook(excel_file)
        sheet = wb["Export"]
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return

    # Find the last row with data
    last_row = sheet.max_row
    print(f"Current last row: {last_row}")

    # Real Weblogic CVEs from 2020-2024 (well-documented vulnerabilities)
    real_weblogic_cves = [
        # 2024 Critical Weblogic CVEs
        {
            "AFFECTED_PLATFORMS": "Windows Server 2022",
            "XTRACT_PATH": "C:\\Oracle\\Middleware\\Oracle_Home\\wlserver\\server\\lib\\weblogic.jar",
            "HOSTNAME": "weblogic-prod-01",
            "CVE": "CVE-2024-20931"  # Weblogic Server vulnerability (Jan 2024)
        },
        {
            "AFFECTED_PLATFORMS": "Linux",
            "XTRACT_PATH": "/opt/oracle/middleware/wlserver/server/lib/weblogic.jar",
            "HOSTNAME": "weblogic-prod-02",
            "CVE": "CVE-2024-21006"  # Weblogic Server vulnerability (Jan 2024)
        },

        # 2023 Critical Weblogic CVEs
        {
            "AFFECTED_PLATFORMS": "Windows Server 2019",
            "XTRACT_PATH": "C:\\Oracle\\Middleware\\wlserver_12.2\\server\\lib\\weblogic.jar",
            "HOSTNAME": "weblogic-test-01",
            "CVE": "CVE-2023-21839"  # Weblogic Server Core vulnerability (Jan 2023)
        },
        {
            "AFFECTED_PLATFORMS": "Linux",
            "XTRACT_PATH": "/u01/oracle/middleware/wlserver/server/lib/weblogic.jar",
            "HOSTNAME": "weblogic-test-02",
            "CVE": "CVE-2023-21931"  # Weblogic Server vulnerability (Apr 2023)
        },
        {
            "AFFECTED_PLATFORMS": "Windows Server 2022",
            "XTRACT_PATH": "C:\\Oracle\\Middleware\\wlserver\\server\\lib\\wls-api.jar",
            "HOSTNAME": "weblogic-dev-01",
            "CVE": "CVE-2023-22067"  # Weblogic Server SAML vulnerability (Oct 2023)
        },

        # 2022 Critical Weblogic CVEs
        {
            "AFFECTED_PLATFORMS": "Linux",
            "XTRACT_PATH": "/opt/weblogic/wlserver/server/lib/weblogic.jar",
            "HOSTNAME": "weblogic-stage-01",
            "CVE": "CVE-2022-21371"  # Weblogic Server vulnerability (Jan 2022)
        },
        {
            "AFFECTED_PLATFORMS": "Windows Server 2019",
            "XTRACT_PATH": "C:\\WebLogic\\wlserver\\server\\lib\\weblogic.jar",
            "HOSTNAME": "weblogic-stage-02",
            "CVE": "CVE-2022-21497"  # Weblogic Server IIOP vulnerability (Apr 2022)
        },
        {
            "AFFECTED_PLATFORMS": "Linux",
            "XTRACT_PATH": "/home/oracle/middleware/wlserver/server/lib/weblogic.jar",
            "HOSTNAME": "weblogic-backup-01",
            "CVE": "CVE-2022-39408"  # Weblogic Server vulnerability (Oct 2022)
        },

        # 2021 Critical Weblogic CVEs
        {
            "AFFECTED_PLATFORMS": "Windows Server 2016",
            "XTRACT_PATH": "C:\\Oracle\\Middleware\\wlserver_12.1\\server\\lib\\weblogic.jar",
            "HOSTNAME": "weblogic-legacy-01",
            "CVE": "CVE-2021-2109"   # Weblogic Server LDAP vulnerability (Jan 2021)
        },
        {
            "AFFECTED_PLATFORMS": "Linux",
            "XTRACT_PATH": "/opt/oracle/wlserver/server/lib/weblogic.jar",
            "HOSTNAME": "weblogic-legacy-02",
            "CVE": "CVE-2021-2394"   # Weblogic Server Core vulnerability (Jul 2021)
        },
        {
            "AFFECTED_PLATFORMS": "Windows Server 2019",
            "XTRACT_PATH": "C:\\Oracle\\wlserver\\server\\lib\\wlthint3client.jar",
            "HOSTNAME": "weblogic-cluster-01",
            "CVE": "CVE-2021-35587"  # Weblogic Server WLS Core vulnerability (Oct 2021)
        },

        # 2020 Critical Weblogic CVEs
        {
            "AFFECTED_PLATFORMS": "Linux",
            "XTRACT_PATH": "/u01/app/oracle/middleware/wlserver/server/lib/weblogic.jar",
            "HOSTNAME": "weblogic-cluster-02",
            "CVE": "CVE-2020-2883"   # Weblogic Server IIOP vulnerability (Apr 2020)
        },
        {
            "AFFECTED_PLATFORMS": "Windows Server 2016",
            "XTRACT_PATH": "C:\\Oracle\\Middleware\\Oracle_Home\\wlserver\\server\\lib\\weblogic.jar",
            "HOSTNAME": "weblogic-dmz-01",
            "CVE": "CVE-2020-14750"  # Weblogic Server vulnerability (Jul 2020)
        },
        {
            "AFFECTED_PLATFORMS": "Linux",
            "XTRACT_PATH": "/opt/bea/wlserver/server/lib/weblogic.jar",
            "HOSTNAME": "weblogic-dmz-02",
            "CVE": "CVE-2020-14756"  # Weblogic Server T3 vulnerability (Jul 2020)
        },
        {
            "AFFECTED_PLATFORMS": "Windows Server 2019",
            "XTRACT_PATH": "C:\\WebLogic\\Oracle_Home\\wlserver\\server\\lib\\weblogic.jar",
            "HOSTNAME": "weblogic-internal-01",
            "CVE": "CVE-2020-14825"  # Weblogic Server Console vulnerability (Oct 2020)
        }
    ]

    # Get header row to find column indices
    headers = {}
    for col in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        if header_value:
            headers[header_value] = col

    print(f"Found headers: {list(headers.keys())}")

    # Add test cases
    for i, test_case in enumerate(real_weblogic_cves):
        row_num = last_row + 1 + i
        print(f"Adding real Weblogic CVE {i+1}/15 to row {row_num}: {test_case['CVE']}")

        for column_name, value in test_case.items():
            if column_name in headers:
                col_num = headers[column_name]
                sheet.cell(row=row_num, column=col_num, value=value)
            else:
                print(f"Warning: Column '{column_name}' not found in headers")

    # Save the file
    try:
        wb.save(excel_file)
        print(f"\nSuccessfully added {len(real_weblogic_cves)} real Weblogic CVEs to {excel_file}")
        print(f"New total rows: {sheet.max_row}")

        print("\n=== REAL WEBLOGIC CVEs ADDED (2020-2024) ===")
        print("2024 CVEs:")
        print("  - CVE-2024-20931: Weblogic Server vulnerability (Critical)")
        print("  - CVE-2024-21006: Weblogic Server vulnerability (High)")

        print("\n2023 CVEs:")
        print("  - CVE-2023-21839: Weblogic Server Core vulnerability (Critical)")
        print("  - CVE-2023-21931: Weblogic Server vulnerability (High)")
        print("  - CVE-2023-22067: Weblogic Server SAML vulnerability (High)")

        print("\n2022 CVEs:")
        print("  - CVE-2022-21371: Weblogic Server vulnerability (High)")
        print("  - CVE-2022-21497: Weblogic Server IIOP vulnerability (High)")
        print("  - CVE-2022-39408: Weblogic Server vulnerability (Critical)")

        print("\n2021 CVEs:")
        print("  - CVE-2021-2109: Weblogic Server LDAP vulnerability (Critical)")
        print("  - CVE-2021-2394: Weblogic Server Core vulnerability (Critical)")
        print("  - CVE-2021-35587: Weblogic Server WLS Core vulnerability (High)")

        print("\n2020 CVEs:")
        print("  - CVE-2020-2883: Weblogic Server IIOP vulnerability (Critical)")
        print("  - CVE-2020-14750: Weblogic Server vulnerability (Critical)")
        print("  - CVE-2020-14756: Weblogic Server T3 vulnerability (Critical)")
        print("  - CVE-2020-14825: Weblogic Server Console vulnerability (Critical)")

        print("\nThese should properly trigger Weblogic detection in the CVE sheet!")

    except Exception as e:
        print(f"Error saving Excel file: {e}")

if __name__ == "__main__":
    add_real_weblogic_cves()